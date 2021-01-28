from time import sleep, perf_counter
from vars import *
import threading
from openpyxl import load_workbook, Workbook
from selenium import webdriver 

def open_website(link):
    """This function just opens the browser on the requested URL"""
    
    ### If you get a "chromedriver.exe must be in PATH" try adding the path in which you have chromedriver installed to
    ### webdriver.Chrome("C:Windows\wherever")
    if chrome_location != "":
        driver = webdriver.Chrome(chrome_location)
    else:
        driver = webdriver.Chrome()
    
    driver.get(link)
    return driver
    
def get_required_info(driver):
    results = []
    ## Get all of the items in the main table
    data_elements = driver.find_elements_by_xpath('//table [@class="table table-condensed table-striped"]/tbody/tr/th')
    ## This loops through all the items and finds the ones we need.
    for i in range(7):
        try:        
            if "Timezone:" in data_elements[i].text:
                results.append(driver.find_elements_by_xpath('//table [@class="table table-condensed table-striped"]/tbody/tr/td')[i].text.split(" ")[0])
            if "Area code:" in data_elements[i].text:
                results.append(driver.find_elements_by_xpath('//table [@class="table table-condensed table-striped"]/tbody/tr/td')[i].text.split(" ")[0])
        except:
            if len(results) < 2:
                results.append("")
    
    ## Looks for items in the secondary tables
    data_elements = driver.find_elements_by_xpath('//table [@class="table table-hover"]/tbody/tr')
    
    ## This loops through the items of the secondary tables to get the remaining items.
    ## It raises an error (to execute the except:) if the item isn't being placed in the correct spot
    i = 0
    while i<8:
        try:    
            if "Population" in data_elements[i].text:
                if "Density" in data_elements[i].text:
                    while len(results) < 3:
                        results.append("")
                        
                    results.append(data_elements[i].text.split(" ")[2]) 
                else:
                    while len(results) < 2:
                        results.append("")
                        
                    results.append(data_elements[i].text.split(" ")[1])
                    
            if "Housing Units" in data_elements[i].text:
                if "Occupied" in data_elements[i].text:
                    while len(results) < 7:
                        results.append("")
                        
                    results.append(data_elements[i].text.split(" ")[3])
                else:
                    while len(results) < 4:
                        results.append("")
                        
                    results.append(data_elements[i].text.split(" ")[2])
            if "Median Home Value" in data_elements[i].text:
                while len(results) < 5:
                    results.append("")
                        
                results.append(data_elements[i].text.split(" ")[3])
            if "Land Area" in data_elements[i].text:
                while len(results) < 6:
                    results.append("")
                        
                results.append(data_elements[i].text.split(" ")[2])
            if "Median Household Income" in data_elements[i].text:
                while len(results) < 8:
                    results.append("")
                        
                results.append(data_elements[i].text.split(" ")[3])
            i += 1
        except:
            ## This should only be reached if the last item gives error
            if len(results) < 9:
                results.append("")
            i += 1


    return results



def thread_operation(start_row, last_row, output_file_name, thread_num):
    """This is the main thread operation, this operates 1 thread. It extracts the zipcodes from zipcode.txt"""
    start = 0
    ## This selects where to resume the scraping  assuming the file is sorted (if the data already exists we do not wanna scrape it.)
    with open ("debugger.txt", "w", encoding = "utf-8") as debugfile:
        debugfile.write("")
    with open(output_file_name, "r", encoding = "utf-8") as file:
        for word in file:
            with open("zipcodes.txt") as zipsfile:
                zipsfile.seek((start_row-1)*7)
                if zipsfile.read(5) == word[0:5]:
                    start_row += 1
    
    
    
    ## This looks at which zipcodes have to be skipped (already scraped...)
    skip_these_zipcodes = []
    with open(output_file_name, "r", encoding = "utf-8") as file:
        with open("zipcodes.txt", "r", encoding = "utf-8") as zipindex:
            zipindex.seek((start_row-1)*7)
            zipstart = int(zipindex.read(5))
            
            zipindex.seek((last_row-1)*7)
            zipend = int(zipindex.read(5))
            
            
        for word in file:
            
            if int(word[:5]) < zipend and int(word[:5]) > zipstart:
                skip_these_zipcodes.append(word[:5])
                
    
    
    sleep(3)
    print("Executed thread", thread_num, "in row", start_row)
    
    
    ## We want to repeat the operation for each zipcode that this thread has to scrape, this is the main operation
    for i in range(start_row, last_row + 1):
        
        ## Just a performance counter, it's accurate. Not correctly implemented but oh well...
        if (i+1) % num_threads == 0:
            start = perf_counter()    
        if i % num_threads == 0:
            print(f"time for {num_threads} zipcodes:", perf_counter() - start)
        
        
        with open("zipcodes.txt") as file:
            file.seek((i-1)*7)
            
            ## This creates the link for the specified zipcode     
            zipcode = file.read(5)
            link = f"https://www.unitedstateszipcodes.org/{zipcode}/"
#         with open(output_file_name, "r", encoding = "utf-8") as file:
            
        if zipcode not in skip_these_zipcodes:
            print("Opening browser for zipcode", zipcode)
            driver = open_website(link)
            
            ## Writes this to the file
            with open(output_file_name, "a", encoding = "utf-8") as file:
                file.write(zipcode + " ")
                file.write(str(get_required_info(driver)))
                file.write("\n")
                print("Saved data for zipcode", zipcode)
            ## Closes browser, hopefully to start a new one.
            driver.close()

def zipcodes_to_txt(worksheet):
    """This function sends the zipcodes to a zipcodes.txt file for easier manipulation"""
    zipcodes = worksheet["A"]
    with open("zipcodes.txt", "w", encoding = "utf-8") as file:
        file.write("-----")
        file.write("\n")
        for i in range(1, len(zipcodes)):
            file.write(zipcodes[i].value)
            file.write("\n")
    print("All zipcodes unpacked")


def clean_txt(file):
    """This function receives an open file removes all entries that are empty
THIS DOES NOT EDIT THE FILE, THIS RETURNS A CLEAN FILE OBJECT THAT CONTAINS ALL LINES ON THE FILE, WITHOUT THE CORRUPTED ONES. OUTER WRITE INSTRUCTION IS REQUIRED"""
    print("Removing corrupted entries")
    memory = []
    file.seek(0)
    j=0
    for line in file:
        if line not in memory:
            memory.append(line)
        else:
            print("Repeated:", line)
            j += 1
        
    # This calculates how many tags there are and how many empty spaces there would be on an empty zipcode entry
    tag_nums = len(tags)
    empty_tags = list("" for i in range(tag_nums))
    
    # This removes corrupted entries
    r = 0
    for i in range(len(memory)):
        try:
            ### This checks if the line doesn't contain empty tags, or if it doesn't follow this format: xxxxx ['...
            # Add (str(empty_tags) in memory[i-r] or) without parentheses to the if, if you wanna remove entries that are empty 
            if memory[i-r][5:8] != " ['":
                print("Corrupted:", memory.pop(i-r))
                r += 1
        except:
            break
    print(f"Removed {r} corrupted entries and {j} repeated entries")
    return memory
        
## sorted_ and clean_txt can be made into a single function...

def sorted_(file):
    """This function receives an open file removes all repeated entries (if there are any) and sorts a text file with information from the zipcodes, the algorithm is bubble sort
THIS DOES NOT EDIT THE FILE, THIS RETURNS A SORTED LIST THAT CONTAINS ALL LINES ON THE FILE, SORTED. OUTER WRITE INSTRUCTION IS REQUIRED"""
    print("Sorting the file (This may take a while)")
    memory =  []
    file.seek(0)
    i = 0
    for line in file:
        memory.append(line)
    for j in range(len(memory)-1):    
        for i in range(len(memory)-j-1):
            if memory[i][:5] > memory[i+1][:5]:
                temp = memory[i]
                memory[i] = memory[i+1]
                memory[i+1] = temp
    print("The file is sorted")
    return memory

def load_spreadsheet():
    """Returns the spreadsheet needed from the template.xlsx file"""
    print("Loading template excel spreadsheet...")
    wbook = load_workbook("template.xlsx")
    wsheet = wbook.active
    return wsheet


def format_data(data):
    """This function gets the string from a file that contains the data and returns TWO VALUES, the first one is
the zipcode, the second is a list with the values."""
    zipcode = data.split(" ")[0]
    zip_data = data.split(" ")[1:]
    clean_list = []
    removethis = ["[", "]", ",", "'", "$"]
    for entry in zip_data:
        for symbol in removethis:
            entry = entry.replace(symbol, "")    
        clean_list.append(entry)

    return zipcode, clean_list


def txt_to_excel(txt_filename, excel_filename, template_worksheet):
    """This procedure receives a the name of a text file with data, an excel name with a template and saves the data to the excel file."""
    print("Updating spreadsheet...")
    with open(txt_filename, "r", encoding = "utf-8") as file:
        row_index = 1
        for line in file:
            zipcode, values = format_data(line)
            while zipcode != template_worksheet.cell(column = 1, row = row_index).value:
                row_index += 1
            col_index = 5
            for item in values:
                template_worksheet.cell(column = col_index, row = row_index, value = item)
                col_index += 1
    print("Spreadsheet updated")



def is_reserved(name):
    if name in reserved_filenames:
        return True
    else:
        return False
    
    
    

def ask_input():
    """REMEMBER: CHANGE NUM_OPT WHEN ADDING AN OPTION
        This function asks the user for input, it's the main menu"""
    ## This is the number of currently supported options
    # When adding a new option, add 1 to this number. The number to choose the option must be the last one + 1 (1. option 1 || 2. option 2... etc.)
    valid = False
    while not valid:
        num_opt = 3
        try:
            ##This is the message shown on the main menu, if you write something that is not a number it will print an error message
            print()
            option = input("""Choose an option from the list and press enter:

1. Resume scraping to an existing text file.
2. Start scraping to a new text file. (If the file already exists, this will overwrite it)
3. Format zipcodes information into an excel file.

Write info for more information

Execute option number: """)
            print()
            print("----------")
            print()
            
            if str(option) == "info":
                """This part is the info menu, here if you choose an option you will get information about it. If you press -1 you go back"""
                print("Enter a number to get information about that option. Enter -1 to go back to the main menu.")
                print()
                inform = input("Information about option number: ")
                print()
                print("----------")
                print()
                
                if inform == "1":
                    print(mode1_doc)
                    print()
                    
                    
                elif inform == "2":
                    print(mode2_doc)
                    print()
                    
                    
                elif inform == "3":
                    print(mode3_doc)
                    print()
                    
                    
                else:
                    """This happens if the user doesn't choose a valid option"""
                    if inform != "-1":
                        raise ValueError
                
                raise TypeError
                
                
                
            else:
                """If the user does not input "info" this make it into an int. It raises a ValueError if it's not one of the options"""
                option = int(option)
                if option < 1 or option > num_opt:
                    raise ValueError
            valid = True
            
        except ValueError:
            if option != -1:
                print("Please enter a valid option.")
                sleep(3)
        except TypeError:
            """We raise a type error to send the user to the main menu again"""
            
            print("Opening the main menu")
            if inform == "-1":
                sleep(3)
            else:
                sleep(10)
    return int(option)

